#!/usr/bin/python3

import requests
import json
import sys
import csv
import time
import datetime as dt
import os
from openpyxl import Workbook
from dotenv import load_dotenv


# TODO what is Check_Current_Workflow_durations

load_dotenv()
ACCESS_TOKEN = os.environ.get("ACCESS_TOKEN")
REPOSITORY = os.environ.get("REPOSITORY")
WORKFLOW = os.environ.get("WORKFLOW")

TIME_FORMAT = "%Y-%m-%dT%H:%M:%SZ"
FILENAME_RAW_JOBS = "{}-{}.json".format(REPOSITORY, WORKFLOW).replace("/", "-")
# FILENAME_WEEKLY_REPORT = "{}-{}-{}.xlsx".format(REPOSITORY, WORKFLOW, dt.datetime.now()).replace("/", "-")
FILENAME_WEEKLY_REPORT = sys.argv[1]


def to_datetime(str_datetime):
    return dt.datetime.strptime(str_datetime, TIME_FORMAT)


def get_last_week_runs():
    now = dt.datetime.now()
    weekly_start_time = now + dt.timedelta(days=-7)
    weekly_runs = []

    url = "https://api.github.com/repos/{}/actions/workflows/{}/runs?page=1&per_page=100".format(
        REPOSITORY, WORKFLOW)
    while True:
        resp = requests.get(url, headers={
            "Authorization": "token {}".format(ACCESS_TOKEN)
        })
        runs = resp.json()["workflow_runs"]
        if len(runs) == 0:
            break

        older_than_last_week = False
        for run in runs:
            if to_datetime(run["created_at"]) < weekly_start_time:
                older_than_last_week = True
                break

            conclusion = run["conclusion"]
            if conclusion == "skipped":
                continue

            status = run["status"]
            if status == "queued":
                continue

            weekly_runs.append(run)

        if older_than_last_week:
            break
        if not ('next' in resp.links.keys()):
            break

        # update the next url
        url = resp.links["next"]["url"]

        print("[DEBUG] {} ~ {} {} runs ...".format(
            weekly_runs[-1]["created_at"], now, len(weekly_runs)))

    return weekly_runs


def get_jobs(run):
    jobs_url = run["jobs_url"]
    resp = requests.get(jobs_url, headers={
        "Authorization": "token {}".format(ACCESS_TOKEN)
    })
    jobs = resp.json()["jobs"]

    jobs = [job for job in jobs if job["completed_at"] is not None]

    # Attach job duratioins onto job fields
    for job in jobs:
        job_name = job["name"]
        job_conclusion = job["conclusion"]
        job_started_at = to_datetime(job["started_at"])
        job_completed_at = to_datetime(job["completed_at"])
        job_durations = (job_completed_at - job_started_at).total_seconds()
        job["durations"] = job_durations

    return jobs


COLUMNS = [
    'A',
    'B',
    'C',
    'D',
    'E',
    'F',
    'G',
    'H',
    'I',
    'J',
    'K',
    'L',
    'M',
    'N',
    'O',
    'P',
    'Q',
    'R',
    'S',
    'T',
    'U',
    'V',
    'W',
    'X',
    'Y',
    'Z',
    'AA',
    'AB',
    'AC',
    'AD',
    'AE',
    'AF']
COLUMNS_MAP = {
    counter: column_name for (
        counter,
        column_name) in enumerate(COLUMNS)}


def column_name(counter):
    return COLUMNS_MAP[counter]


def column_title_position(counter):
    return "{}1".format(COLUMNS_MAP[counter])


def report(runs):
    wb = Workbook()

    # align sheet "WorkflowRuns"
    ws_runs = wb.active
    ws_runs.title = "WorkflowRuns"
    align_sheet_workflow_runs(ws_runs)

    # align sheet "WorkflowJobs"
    ws_jobs = wb.create_sheet("WorkflowJobs")
    ws_jobs[column_title_position(0)] = "Workflow ID"
    jobs_titles = get_jobs_titles(runs)
    for (i, title) in enumerate(jobs_titles):
        ws_jobs[column_title_position(i + 1)] = title
    title_to_column = {}
    for (i, title) in enumerate(jobs_titles):
        title_to_column[title] = column_name(i + 1)

    columns_with_durations = [title_to_column[title]
                              for title in jobs_titles if "duration" in title]

    for run in runs:
        run_workflow_id = run["workflow_id"]
        run_author = run["head_commit"]["author"]["name"]
        run_created_at = to_datetime(run["created_at"])
        run_updated_at = to_datetime(run["updated_at"])
        run_durations = (run_updated_at - run_created_at).total_seconds()
        jobs = run["jobs"]

        # append jobs to sheet "WorkflowJobs"
        job_metric = {}
        for job in jobs:
            job_name = job["name"]
            conclusion = job["conclusion"]
            durations = job["durations"]
            job_metric["{} conclusion".format(job_name)] = conclusion
            job_metric["{} durations".format(job_name)] = durations

        if not all([title in job_metric for title in jobs_titles]):
            continue

        jobs_raw_metrics = []
        jobs_raw_metrics.append(run_workflow_id)
        for title in jobs_titles:
            value = job_metric[title]
            jobs_raw_metrics.append(value)

        ws_jobs.append(jobs_raw_metrics)

        if len(jobs) == 0:
            continue

        # append workflows to sheet "WorkflowRuns"
        pending_time = (
            to_datetime(
                jobs[0]["started_at"]) -
            run_created_at).total_seconds()
        duration_jobs_cols = ",".join(["WorkflowJobs!{}{}".format(
            col, ws_jobs.max_row) for col in columns_with_durations])
        avg_jobs_execution_time = "=AVERAGE({})".format(duration_jobs_cols)
        max_jobs_execution_time = "=MAX({})".format(duration_jobs_cols)
        sum_jobs_execution_time = "=SUM({})".format(duration_jobs_cols)
        ws_runs.append([
            run_workflow_id,
            run_created_at,
            run["event"],
            run_author,
            run["head_branch"],
            run["conclusion"],
            run_durations,
            pending_time,
            avg_jobs_execution_time,
            max_jobs_execution_time,
            sum_jobs_execution_time
        ])

    # sheet DataAnalyze
    workflow_run_avg = '=AVERAGE(WorkflowRuns!G2:G' + \
        str(ws_runs.max_row) + ')/' + str(60)
    workflow_success_percentile_99 = '=PERCENTILE(IF(WorkflowRuns!F2:F' + str(
        ws_runs.max_row) + '="success",WorkflowRuns!G2:G' + str(ws_runs.max_row) + '),0.99)/' + str(60)
    workflow_max = '=MAX(WorkflowRuns!G2:G' + \
        str(ws_runs.max_row) + ')/' + str(60)
    workflow_success_rate = '=COUNTIF(WorkflowRuns!F2:F' + str(
        ws_runs.max_row) + ',"success")/' + str(ws_runs.max_row - 1) + '*100'

    workflow_wait_avg = '=AVERAGE(WorkflowRuns!H2:H' + \
        str(ws_runs.max_row) + ')/' + str(60)
    workflow_success_wait_percentile_99 = '=PERCENTILE(IF(WorkflowRuns!F2:F' + str(
        ws_runs.max_row) + '="success",WorkflowRuns!H2:H' + str(ws_runs.max_row) + '),0.99)/' + str(60)
    workflow_wait_max = '=MAX(WorkflowRuns!H2:H' + \
        str(ws_runs.max_row) + ')/' + str(60)

    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = "workflow_run_avg(minues)"
    ws_summary['B1'] = workflow_run_avg
    ws_summary['A2'] = "workflow_success_percentile_99(minues)"
    ws_summary['B2'] = workflow_success_percentile_99
    ws_summary['A3'] = "workflow_max(minues)"
    ws_summary['B3'] = workflow_max
    ws_summary['A4'] = "workflow_success_rate(%)"
    ws_summary['B4'] = workflow_success_rate

    ws_summary['A6'] = "workflow_wait_avg(minues)"
    ws_summary['B6'] = workflow_wait_avg
    ws_summary['A7'] = "workflow_success_wait_percentile_99(minues)"
    ws_summary['B7'] = workflow_success_wait_percentile_99
    ws_summary['A8'] = "workflow_wait_max(minues)"
    ws_summary['B8'] = workflow_wait_max

    for i, title in enumerate(jobs_titles):
        if not ("durations" in title):
            continue

        title_job_durations = title
        title_job_conclusion = jobs_titles[i - 1]
        column_name_job_durations = title_to_column[title_job_durations]
        column_name_job_conclusion = title_to_column[title_job_conclusion]

        i_row = ws_summary.max_row + 1
        ws_summary["A{}".format(
            i_row + 1)] = "Avg {} (minutes)".format(title_job_durations)
        ws_summary["B{}".format(i_row + 1)] = "=AVERAGE(WorkflowJobs!{}2:{}{})/60".format(
            column_name_job_durations, column_name_job_durations, ws_jobs.max_row)
        ws_summary["A{}".format(
            i_row + 2)] = "Percentile-99 {} (minutes)".format(title_job_durations)
        ws_summary["B{}".format(i_row + 2)] = "=PERCENTILE(IF(WorkflowJobs!{}2:{}{}=\"success\",WorkflowJobs!{}2:{}{}),0.99)/60".format(
            column_name_job_conclusion, column_name_job_conclusion, ws_jobs.max_row, column_name_job_durations, column_name_job_durations, ws_jobs.max_row)
        ws_summary["A{}".format(
            i_row + 3)] = "Max {} (minutes)".format(title_job_durations)
        ws_summary["B{}".format(i_row + 3)] = "=Max(WorkflowJobs!{}2:{}{})/60".format(
            column_name_job_durations, column_name_job_durations, ws_jobs.max_row)
        ws_summary["A{}".format(
            i_row + 4)] = "Success Rate of {} (minutes)".format(title_job_durations)
        ws_summary["B{}".format(i_row + 4)] = "=COUNTIF(WorkflowJobs!{}2:{}{}, \"success\")/{}*100".format(
            column_name_job_conclusion, column_name_job_conclusion, ws_jobs.max_row, ws_jobs.max_row)

    wb.save(FILENAME_WEEKLY_REPORT)


# A sheet lists workflow runs
def align_sheet_workflow_runs(ws):
    ws['A1'] = "ID"
    ws['B1'] = "Created"
    ws['C1'] = "Event"
    ws['D1'] = "Author"
    ws['E1'] = "Branch"
    ws['F1'] = "Conclusion"
    ws['G1'] = "Durations"
    ws['H1'] = "Pending Time"
    ws['I1'] = "Avg Jobs Execution Time"
    ws['J1'] = "Max Jobs Execution Time"
    ws['K1'] = "Sum Jobs Execution Time"


def get_jobs_titles(runs):
    titles = set([])
    for run in runs:
        jobs = run["jobs"]
        for job in jobs:
            name = job["name"]
            durations = "{} durations".format(name)
            conclusion = "{} conclusion".format(name)
            titles.add(durations)
            titles.add(conclusion)
    return [title for title in sorted(titles)]


def main():
    weekly_runs = get_last_week_runs()
    for run in weekly_runs:
        run["jobs"] = get_jobs(run)

    with open(FILENAME_RAW_JOBS, "w") as jobs_f:
        json.dump(weekly_runs, jobs_f)

    with open(FILENAME_RAW_JOBS, "r") as jobs_f:
        weekly_runs = json.load(jobs_f)
        report(weekly_runs)


if __name__ == '__main__':
    main()
